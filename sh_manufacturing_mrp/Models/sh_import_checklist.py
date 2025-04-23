from odoo import fields,models,api
import csv,io
import importlib,base64,xlrd,xlsxwriter
from openpyxl import load_workbook
from odoo.exceptions import UserError, ValidationError


# import pandas as pd

class shMrpImportExportChecklist(models.TransientModel):

    _name = "sh.import.export.checklist"
    _description = "import export checklist"
    
    
    name = fields.Selection(
        string='import file type',
        selection=[('csv', 'CSV File'), ('excel', 'Excel File')],default='csv'
    )

    
    company_id = fields.Many2one(
        string='Company', 
        comodel_name='res.company', 
        required=True, 
        default=lambda self: self.env.company.id
    )

    file = fields.Binary(string='File')
    file_name = fields.Char()

    def checklist_import_export_apply(self):
        # print(f"\n\n\n\t--------------> 27 ",self.file,"\n\nType",type(self.file))
        self.process_file()

    def process_file(self):
            if self.file:
                if self.name == "csv":
                    self._process_csv()
                elif self.name == "excel":
                    self._process_excel()
                else:
                    raise UserError("Unsupported file format. Please upload CSV or Excel file.")
            else:
                raise UserError("Please enter the file")
 
    def _process_csv(self):
        decoded_file = base64.b64decode(self.file)
 
        try:
            file_io = io.StringIO(decoded_file.decode("utf-8"))
            reader = csv.DictReader(file_io)
        except Exception as e:
            raise UserError(f"Failed to read CSV content: {str(e)}")
 
        for row in reader:
            # Example: Create a partner
            # print('\n\n\n-----row.get("name")------->',row.get("Name"))
            # print('\n\n\n-----row.get("name")------->',row.get("Description"))
            self.env["sh.manufacturing.checklist"].create(
                {
                    "name": row.get("Name"),
                    "description": row.get("Description"),
                    "company_id": self.company_id.id,
                }
            )
 
    def _process_excel(self):
        decoded_file = base64.b64decode(self.file)

        try:
            file_io = io.BytesIO(decoded_file)
            workbook = load_workbook(filename=file_io, read_only=True)
            sheet = workbook.active

            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))

                # print(f"Row: {data}")

                self.env["sh.manufacturing.checklist"].create({
                    "name": data.get("Name"),
                    "description": data.get("Description"),
                    "company_id": self.company_id.id,
                })

        except Exception as e:
            raise UserError(f"Failed to read Excel file: {str(e)}")

    
    def checklist_export_csv(self):
        # print(f"\n\n\n\t--------------> 30 CSV ")
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['Name', 'Description'])
        writer.writerow(['Cheklist A', 'This is description about checklist A'])
        writer.writerow(['Cheklist B', 'This is description about checklist B'])

        csv_buffer.seek(0)
        data = csv_buffer.read()
        csv_buffer.close()

        # Encode to base64
        csv_base64 = base64.b64encode(data.encode('utf-8'))

        # Create an attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'sample.csv',
            'type': 'binary',
            'datas': csv_base64,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv'
        })

        # Generate URL for download
        download_url = f'/web/content/{attachment.id}?download=true'

        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }

    def checklist_export_excel(self):
        # print(f"\n\n\n\t--------------> 33 Excel",)

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Sample data
        headers = ['Name', 'Description']
        data = [
            ['Cheklist A', 'This is description about checklist A'],
            ['Cheklist B', 'This is description about checklist B'],
        ]


        # Write headers
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # Write data rows
        for row_num, row_data in enumerate(data, start=1):
            for col_num, cell in enumerate(row_data):
                worksheet.write(row_num, col_num, cell)

        workbook.close()
        output.seek(0)

        # Encode Excel content to base64
        excel_base64 = base64.b64encode(output.read())

        # Create the attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'sample.xlsx',
            'type': 'binary',
            'datas': excel_base64,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Generate download URL
        download_url = f'/web/content/{attachment.id}?download=true'

        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }
        

# class ShContact(models.Model):
#     _inherit="res.partner"
    
#     company_type = fields.Selection(string='Company Type',
#         selection=[('person', 'Patient'), ('company', 'Doctor')],
#         compute='_compute_company_type', inverse='_write_company_type')